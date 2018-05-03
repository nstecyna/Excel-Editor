import openpyxl
import search
import changeValue

def searchMenu(filename, wb):
    '''
    Asks for the excel filename and continuously asks for the next action,
    until the user inputs "exit".
    param list:
        (str) filename = the filename of the active excel document
        (Workbook) wb = a workbook of the active excel document
    '''

    print()

    # only moves forward when sheetIn is one of the sheets in the wb
    sheetIn = ""
    while sheetIn not in wb.sheetnames:
        sheetIn = input("Please input a valid excel sheet in this document "
            + "(case-sensitive) " + str(wb.sheetnames) + ": ")
    sheet = wb[sheetIn]

    itemCol = input("Please input the column for the item names: ").upper()
    amtCol = input("Please input the column for the current inventory: ").upper()

    cell = search.search(sheet, itemCol)

    inSearchMenu = True
    while (inSearchMenu):
        print("\nCurrent Search: " + str(wb[sheetIn][cell.column][cell.row - 1].value)
            + ": " + str(wb[sheetIn][amtCol][cell.row - 1].value))
        print("What would you like to do with this item?")
        choice = input("You can choose to: New Search; Change Name; Change Amount; Exit  \n")
        if (choice.lower() in ("new search", "newsearch", "new", "search", "n s", "ns", "s")):
            cell = search.search(wb[sheetIn], itemCol)
        elif (choice.lower() in ("change name", "name change", "changename", "namechange", "change", "name", "c n", "cn", "n")):
            # can't use cell for the cell parameter because this causes this
            # menu to not update with the correct new name, even though the
            # name will change on the document
            changeValue.changeValue(filename, wb, sheetIn, wb[sheetIn][cell.column][cell.row - 1])
            wb = openpyxl.load_workbook(filename)
        elif (choice.lower() in ("change amount", "amount change", "changeamount", "amountchange", "change", "amount", "c a", "ca", "a")):
            changeValue.changeValue(filename, wb, sheetIn, wb[sheetIn][amtCol][cell.row-1])
            wb = openpyxl.load_workbook(filename)
        elif (choice.lower() in ("exit", "e")):
            inSearchMenu = False
        else:
            print("Invalid choice.")
